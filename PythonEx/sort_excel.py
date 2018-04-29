#!/usr/bin/env python
# -*- coding: utf-8 -*-

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# 读取表格
wb_sum = load_workbook('F:\\3天考核-4.17-4.19\\3天考核汇总.xlsx')  # 总表
sheet_sum = wb_sum.worksheets[0]

wb_overdue1 = load_workbook('F:\\3天考核-4.17-4.19\\c-m1逾期率day1.xlsx')  # 逾期表1
sheet_overdue1 = wb_overdue1.worksheets[0]
wb_overdue2 = load_workbook('F:\\3天考核-4.17-4.19\\c-m1逾期率day2.xlsx')  # 逾期表2
sheet_overdue2 = wb_overdue2.worksheets[0]
wb_overdue3 = load_workbook('F:\\3天考核-4.17-4.19\\c-m1逾期率day3.xlsx')  # 逾期表3
sheet_overdue3 = wb_overdue3.worksheets[0]

wb_finish1 = load_workbook('F:\\3天考核-4.17-4.19\\放款达成率day1.xlsx')  # 达成表1
sheet_finish1 = wb_finish1.worksheets[0]
wb_finish2 = load_workbook('F:\\3天考核-4.17-4.19\\放款达成率day2.xlsx')  # 达成表2
sheet_finish2 = wb_finish2.worksheets[0]
wb_finish3 = load_workbook('F:\\3天考核-4.17-4.19\\放款达成率day3.xlsx')  # 达成表3
sheet_finish3 = wb_finish3.worksheets[0]


def get_sheet_info(sheet, search_data, search_col, target_col):
    """
    查询表格获取对应行的指定数据
    :param sheet: 要查询的表格
    :param search_data: 要对比的属性内容
    :param search_col: 要对比的属性所在列，用字母表示，如A
    :param target_col: 要查找的属性所在列，用字母表示，如A
    :return: 查找的属性值。-1表示无值，-2表示该公司只有平均值
    """

    value = -1
    nrows = sheet.max_row
    # 遍历比较
    for i in range(1, nrows + 1):  # openpyxl的行列从1开始
        index_search = search_col + str(i)  # 单元格编号，如A4
        tp_data = sheet[index_search].value
        if search_data == tp_data:
            date = sheet['C' + str(i)].value  # 日期判断仅针对逾期表
            if date != '平均':
                index_value = target_col + str(i)
                value = sheet[index_value].value
                if isinstance(value, str) and '%' in value:  # 判断是否为字符串形式的百分比
                    value = float(value[:-1]) / 100
            elif value < 0:
                print('该分公司仅有平均值')
                value = -2
    return value


# 遍历逾期表1，匹配分行名称，获取C->M1（单数迁移率）
n_of_rows = sheet_sum.max_row  # 总行数
cp_col = 'B'  # 分行名所在列编号
# 遍历查询所有分行
for i in range(3, n_of_rows):  # openpyxl的行列从1开始
    name = sheet_sum[cp_col + str(i)].value
    print('分行名 ： %s' % name)

    # 获取day1的C-M1入催率
    overdue1_value = get_sheet_info(sheet_overdue1, name + '分公司', 'B', 'H')
    if overdue1_value >= 0:
        tp_value = str(round(overdue1_value * 100,2)) + '%'
        sheet_sum['C' + str(i)].value = tp_value
        print('day1 C-M1入催率：%s' % tp_value)
    elif overdue1_value == -2:
        sheet_sum['C' + str(i)].value = '平均'

    # 获取day2的C-M1入催率
    overdue2_value = get_sheet_info(sheet_overdue2, name + '分公司', 'B', 'H')
    if overdue2_value >= 0:
        tp_value = str(round(overdue2_value * 100,2)) + '%'
        sheet_sum['D' + str(i)].value = tp_value
        print('day2 C-M1入催率：%s' % tp_value)
    elif overdue2_value == -2:
        sheet_sum['D' + str(i)].value = '平均'

    # 获取day3的C-M1入催率
    overdue3_value = get_sheet_info(sheet_overdue3, name + '分公司', 'B', 'H')
    if overdue3_value >= 0:
        tp_value = str(round(overdue3_value * 100,2)) + '%'
        sheet_sum['E' + str(i)].value = tp_value
        print('day3 C-M1入催率：%s' % tp_value)
    elif overdue3_value == -2:
        sheet_sum['E' + str(i)].value = '平均'

    tp_num = sheet_sum['L' + str(i)].value  # 放款目标，单位万
    # 获取day1的放款达成
    finish1_value = get_sheet_info(sheet_finish1, name + '分公司', 'A', 'C')
    if finish1_value >= 0:
        if tp_num > 0:
            finish1_value = finish1_value / 10000 / tp_num  # 原值为类277000整数，故需先计算百分比
            tp_value = str(round(finish1_value * 100, 2)) + '%'
        else:
            tp_value = '0.00%'
        sheet_sum['F' + str(i)].value = tp_value
        print('day1 放款达成率：%s' % tp_value)

    # 获取day2的放款达成
    finish2_value = get_sheet_info(sheet_finish2, name + '分公司', 'A', 'C')
    if finish2_value >= 0:
        if tp_num > 0:
            finish2_value = finish2_value / 10000 / tp_num  # 原值为类277000整数，故需先计算百分比
            tp_value = str(round(finish2_value * 100, 2)) + '%'
        else:
            tp_value = '0.00%'
        sheet_sum['G' + str(i)].value = tp_value
        print('day2 放款达成率：%s' % tp_value)

    # 获取day3的放款达成
    finish3_value = get_sheet_info(sheet_finish3, name + '分公司', 'A', 'C')
    if finish3_value >= 0:
        if tp_num > 0:
            finish3_value = finish3_value / 10000 / tp_num  # 原值为类277000整数，故需先计算百分比
            tp_value = str(round(finish3_value * 100, 2)) + '%'
        else:
            tp_value = '0.00%'
        sheet_sum['H' + str(i)].value = tp_value
        print('day3 放款达成率：%s' % tp_value)

    # 判断两类是否达标
    ncm1 = 0  # c-m1达票次数
    if overdue1_value >= 0 and overdue1_value <= 0.1:  # 此时overdue1_value值应为类0.0453数据
        ncm1 += 1
    if overdue2_value >= 0 and overdue2_value <= 0.1:
        ncm1 += 1
    if overdue3_value >= 0 and overdue3_value <= 0.1:
        ncm1 += 1
    sheet_sum['J' + str(i)].value = ncm1

    nfinish = 0  # 放款达标次数
    if finish1_value >= 1:
        nfinish += 1
    if finish2_value >= 1:
        nfinish += 1
    if finish3_value >= 1:
        nfinish += 1
    sheet_sum['K' + str(i)].value = nfinish

    # 判断总体是否达标
    if ncm1 > 0 or nfinish > 0:
        sheet_sum['I' + str(i)].value = '是'
    else:
        sheet_sum['I' + str(i)].value = '否'

wb_sum.save('F:\\a.xlsx')
