#!/usr/bin/env python
# -*- coding: utf-8 -*-

from openpyxl import load_workbook
from openpyxl import Workbook

wb_source = load_workbook('F:\\M1-三天数据处理过程.xlsx')  # 总表
sheet_source = wb_source.worksheets[3]

wb_target=Workbook()
sheet_target=wb_target.get_active_sheet()

ital=['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']

"""
获取excel表格下一列的编号，如当前列为A，下一列则为B
:param now_col: 当前列
:return: 下一列
"""
def get_next_excel_col(col_now):
    if len(col_now)==1:
        i=ital.index(col_now)
        if i<=24:
            col_next=ital[i+1]
        else:
            col_next='AA'
    else:
        first=col_now[0]
        second=col_now[1]
        i=ital.index(second)
        if i<=24:
            col_next=first+ital[i+1]
        else:
            j=ital.index(first)
            col_next=ital[j+1]+'A'
    return col_next

# 遍历表格，获取exp,blank和after的数据列表
col_exp= 'A'
col_blank= 'M'
col_after= 'X'
deviation=0.005
nrows=sheet_source.max_row
ncol=sheet_source.max_column
exps=[]
blanks=[]
afters=[]

sheet_target[col_exp+'1']='Sample-exp'
for i in range(2,nrows+1):
    col_now = col_exp   # 当前列，从exp开始
    num_exp=sheet_source[col_exp + str(i)].value #当前行exp数字

    if not num_exp:
        break
    sheet_target[col_exp + str(i)]=num_exp
    # 获取exp计算结果
    col_now=get_next_excel_col(col_now)
    while col_now!=col_blank and sheet_source[col_now + '1'].value:   # 当未遍历至blank列或空列继续循环
        # 获取本列属性名及当前值
        key_exp=sheet_source[col_now + '1'].value
        if not key_exp:
            break
        value_exp=sheet_source[col_now + str(i)].value
        if not value_exp:
            value_exp=0
        sheet_target[col_now + '1'] = key_exp   # 写入新表列名

        value_sub=0     # 要被减去的blank及after值
        num_sub=0       # 要被减去的blank及after行的数量，用以计算均值
        # 遍历blank及after的所有行，找出偏差范围（0.005）内的数据,计算差值
        for j in range(2,nrows+1):
            # 计算blank区域数据
            col_tp=col_blank
            num_blank=sheet_source[col_blank + str(j)].value # 当前行blank数字
            # 在满足偏差的列继续查询
            if num_blank and abs(num_blank-num_exp)<=deviation:
                col_tp = get_next_excel_col(col_tp)
                while col_tp != col_after and sheet_source[col_tp + str(j)].value:  # 当未遍历至after列或空列时继续循环
                    # 获取本列属性值
                    key_blank = sheet_source[col_tp + '1'].value  # 本列blank属性名
                    value_blank = sheet_source[col_tp + str(j)].value
                    # 当为同一属性且值存在时，累加计算
                    if key_blank==key_exp and value_blank:
                        value_sub+=value_blank
                        num_sub+=1
                        break
                    else:
                        col_tp=get_next_excel_col(col_tp)

            # 计算after区域数据
            col_tp=col_after
            num_after=sheet_source[col_after + str(j)].value # 当前行after数字
            # 在满足偏差的列继续查询
            if num_after and abs(num_after-num_exp)<=deviation:
                col_tp = get_next_excel_col(col_tp)
                while sheet_source[col_tp + str(j)].value:  # 当未遍历至空列时继续循环
                    # 获取本列属性值
                    key_after = sheet_source[col_tp + '1'].value  # 本列after属性名
                    value_after = sheet_source[col_tp + str(j)].value
                    # 当为同一属性且值存在时，累加计算
                    if key_after==key_exp and value_after:
                        value_sub+=value_after
                        num_sub+=1
                        break
                    else:
                        col_tp=get_next_excel_col(col_tp)
        # 计算累加值的均值
        if num_sub>1:
            value_sub/=num_sub

        value_end=value_exp-value_sub
        print('[{}/{}] key:{}\texp:{}\tsub:{}\tend:{}'.format(i,nrows,key_exp,value_exp,value_sub,value_end))
        # 在当前位置写入计算后的数据
        sheet_target[col_now + str(i)]=value_end
        col_now = get_next_excel_col(col_now)

wb_target.save(r'F:\result.xlsx')