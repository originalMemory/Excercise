#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
import re
import pymysql.cursors
from openpyxl import load_workbook
from urllib import request
from urllib import parse
import chardet

def get_bool(str):
    '''
    判断字符串的布尔值
    :param str: 被判断的字符串
    :return: 布尔值
    '''
    if str == '' or str == 'False':
        return False
    else:
        return True

def get_html(url):
    """
    获取html页面
    :param url:网页链接
    :return: 解析过的网页源码
    """

    head = dict()
    # 写入User Agent信息
    head[
        'User-Agent'] = 'Mozilla/5.0 (Windows NT 6.2; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) ' \
                        'Chrome/27.0.1453.94 Safari/537.36'
    # 创建Request对象
    req = request.Request(url, headers=head)

    # 传入创建好的Request对象
    response = request.urlopen(req)

    # 读取响应信息并解码
    tp_html = response.read()
    charset = chardet.detect(tp_html)
    tp_html = tp_html.decode(charset['encoding'])
    return tp_html

# mysql配置参数
config = {
    'host': 'localhost',  # 地址
    'user': 'root',  # 用户名
    'passwd': 'kdi1994',  # 密码
    'db': 'myacgn_info',  # 使用的数据库名
    'charset': 'utf8',  # 编码类型
    'cursorclass': pymysql.cursors.DictCursor  # 按字典输出
}

def sort_manga(file_path):
    '''
    整理漫画，写入mysql
    :param file_path: 文件名
    :return:
    '''

    db = pymysql.connect(**config)
    cursor = db.cursor()

    wb = load_workbook('F:\\comic.xlsx')
    sheet = wb.worksheets[0]

    # 遍历第一行，确定各属性位置
    series_i = 'A'  # 同人作品原作名
    title_i = 'B'  # 标题
    author_i = 'C'  # 作者
    publisher_i = 'D'  # 出版地
    is_tankoubon_i = 'E'  # 是否是单行本
    size_i = 'F'  # 大小
    cover_i = 'G'  # 封面
    desc_i = 'H'  # 描述
    is_mushuusei_i = 'I'  # 是否无修
    add_date_i = 'J'  # 添加日期
    tag_i = 'K'  # 标签
    rating_i = 'M'  # 评分（1~5）
    lan_i = 'N'  # 语言
    is_read_i = 'O'  # 是否已读
    tran_title_i = 'P'  # 汉化名

    nrow = sheet.max_row
    for i in range(2, nrow + 1):
        comic = {
            'title': '',
            'authors': '',
            'tran_title': '',
            'cover_path': '',
            'tags': '',
            'description': '',
            'size': 0,
            'is_readed': False,
            'rating': 0,
            'is_mushuusei': False,
            'is_tankoubon': False,
            'is_color': False,
            'original_work': '',
            'create_time': '0000-00-00',
            'laguage': '中文',
            'publisher': '',
        }
        if i == 965:
            df = 34
        comic['title'] = sheet[title_i + str(i)].value
        if "'" in comic['title']:
            comic['title'] = comic['title'].replace("'", "\\'")
        comic['authors'] = str(sheet[author_i + str(i)].value)
        if "'" in comic['authors']:
            comic['authors'] = comic['authors'].replace("'", "\\'")
        comic['tran_title'] = sheet[tran_title_i + str(i)].value
        comic['cover_path'] = sheet[cover_i + str(i)].value
        if "'" in comic['cover_path']:
            comic['cover_path'] = comic['cover_path'].replace("'", "\\'")
        comic['tags'] = sheet[tag_i + str(i)].value
        comic['description'] = sheet[desc_i + str(i)].value
        if comic['description'] and "'" in comic['description']:
            comic['description'] = comic['description'].replace("'", "\\'")
        size = float(sheet[size_i + str(i)].value)
        size = size / 1024 / 1024
        size = round(size, 2)
        comic['size'] = size
        comic['is_readed'] = get_bool(sheet[is_read_i + str(i)].value)
        if sheet[rating_i + str(i)].value is None:
            comic['rating'] = 0
        else:
            comic['rating'] = int(sheet[rating_i + str(i)].value)
        comic['is_mushuusei'] = get_bool(sheet[is_mushuusei_i + str(i)].value)
        comic['is_tankoubon'] = get_bool(sheet[is_tankoubon_i + str(i)].value)
        if comic['tags'] and '全彩' in comic['tags']:
            comic['is_color'] = True
        comic['original_work'] = sheet[series_i + str(i)].value
        if sheet[lan_i + str(i)].value == 'jpn':
            comic['laguage'] = '日语'
        comic['publisher'] = sheet[publisher_i + str(i)].value
        comic['create_time'] = sheet[add_date_i + str(i)].value.replace('T', ' ').replace('+08:00', '')

        info = '[{}/{}] {}'.format(i, nrow, comic['title'])
        if comic['tran_title']:
            info += '/{}'.format(comic['tran_title'])
        print(info)

        sql = "select title from comic where title like '{}'".format(comic['title'])
        cursor.execute(sql)
        query = cursor.fetchone()
        if i == 1830:
            df = 34
        if not query:
            sql = "insert into comic(title,authors,tran_title,cover_path,tags,description,size,is_readed,rating," \
                  "is_mushuusei,is_tankoubon,is_color,original_work,laguage,publisher,create_time) values(" \
                  "'{}','{}','{}','{}','{}','{}',{},{},{},{},{},{},'{}','{}','{}','{}');".format(
                comic['title'], comic['authors'], comic['tran_title'], comic['cover_path'], comic['tags']
                , comic['description'], comic['size'], comic['is_readed'], comic['rating'], comic['is_mushuusei']
                , comic['is_tankoubon'], comic['is_color'], comic['original_work'], comic['laguage'], comic['publisher']
                , comic['create_time'])
            sql = sql.replace('None', '')
            cursor.execute(sql)
            db.commit()
    cursor.close()
    db.close()


def sort_txt(txt_dir,target_file):
    '''
    整理飞卢小说
    :param txt_dir: 分段文本所在文件夹
    :param target_file: 整理合并后的文本路径
    :return:
    '''
    # 加载关键词转换
    error_word = []
    with open('error_word.txt', 'r', encoding='utf-8') as f:
        word = f.readline()
        while word:
            error_word.append(word.replace('\n','').split(' '))
            word = f.readline()

    # 获取所有小说文件
    txt_files=[]
    for dir_path, dir_names, file_names in os.walk(txt_dir):
        for file in file_names:
            if not re.search(r'.txt$|.TXT$',file):
                continue
            else:
                txt_files.append(os.path.join(dir_path,file))

    new_txt = []
    text = ""
    pargraph = ''  # 去除非段落换行后的一段
    for file in txt_files:
        with open(file, 'r') as f:
            text = f.readlines()
        for txt in text:
            if len(txt)==0:
                continue
            if '本章内容包含图片，点击屏幕右下角“插图”按钮查看图片。' in txt:
                print(txt)
                continue
            if '求鲜花' in txt:
                print(txt)
                continue
            if re.match(r'我的二次元人生', txt):
                continue

            txt=txt.strip()
            # 判断是否为页面底说明
            # 2，6 096章数字谜题    22:18厦i
            mat = re.search(r'^\d?.\d.{1,2}[第]?[0-9一二三四五六七八九〇零十百千万]{2,5}', txt)
            if mat:
                print(txt)
                continue

            # 判断该行是否为一句话的终点且未能识别出‘”’
            mat = re.search(r'^“', txt)
            if mat:
                mat2 = re.search(r'“$', txt)
                if mat2:
                   txt=re.sub(r'“$','”', txt)
            # 判断以‘道’结束时是否未识别出‘：’
            txt = re.sub(r'道$', '道：', txt)
            mat = re.match(r'第?[0-9一二三四五六七八九〇零十百千万]{2,5}章?', txt)
            if mat:
                print(mat.group())
                txt = txt.replace(mat.group(), mat.group() + ' ')
                # txt=txt.replace('第','')
            for error in error_word:
                txt = txt.replace(error[0], error[1])
            pargraph += txt
            # 判断是否已经到了一句的结尾或者是章节名
            if re.search('。$|？$|”$|！$|）$|：$|>$',txt) or re.search('^[第]?[0-9一二三四五六七八九〇零十百千万]{3,5}.{0,2}[章|节]?',txt):
                new_txt.append(pargraph)
                pargraph=''

    with open(target_file, 'w+') as f:
        for txt in new_txt:
            f.write(txt+'\n')
    print('处理结束')

def img_cut_tran(source_dir,target_dir=None, target_ex=None, left=None, top=None, right=None, bottom=None):
    '''
    批量图片转换及裁剪
    :param source_dir: 图片所在原文件夹
    :param target_dir: 保存文件夹
    :param target_ex: 目标格式后缀（.jpg）
    :param left:裁剪左边缘（为空时不裁剪）
    :param top: 裁剪上边缘
    :param right: 裁剪右边缘
    :param bottom: 裁剪底边缘
    :return:
    '''

    import PIL.Image as Image

    # 获取文件夹内所有文件，只保留图片
    files=[]
    for name in os.listdir(source_dir):
        if name.endswith('.jpg') or name.endswith('.png') or name.endswith('.bmp') or name.endswith('.jpeg') or name.endswith('.tif'):
            files.append(os.path.join(source_dir, name))

    if not target_dir:
        target_dir=source_dir

    i=1
    for file in files:
        filename=os.path.basename(file)
        print("[{}/{}] {}".format(i,len(files),filename))
        i+=1
        # 读取图片
        img = Image.open(file)
        # 裁剪
        if left!=None:
            box = [left, top, right, bottom]
            img = img.crop(box)
        ex=os.path.splitext(filename)[1]
        if target_ex:
            filename = filename.replace(ex, target_ex)
        if target_ex=='.jpg':
            img=img.convert('RGB')
        new_file=os.path.join(target_dir,filename)
        img.save(new_file)