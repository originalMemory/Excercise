#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
import re
import pymysql.cursors
import time
from openpyxl import load_workbook


def get_bool(str):
    '''
    判断字符串的布尔值
    :param str: 被判断的字符串
    :return: 布尔值
    '''
    if str == '' or str == 'False':
        return False
    else:
        return True


letter = ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V',
          'W', 'X', 'Y', 'Z')


def get_col_letter(col_index):
    """
    获取列数字序号对应的字母序号（仅至双字母编号，无法计算三字母编号）
    :param col_index: 列数字序号（从0开始）
    :return: 列字母序号
    """
    # 判断序号是否已超过AA
    if col_index <= 25:
        return letter[col_index]
    else:
        # 判断首字母序号
        initial = letter[int(col_index / 26) - 1]
        # 获取后续字母序号
        second = letter[col_index % 26]
        return initial + second


config = {
    'host': 'localhost',  # 地址
    'user': 'root',  # 用户名
    'passwd': 'kdi1994',  # 密码
    'db': 'myacgn_info',  # 使用的数据库名
    'charset': 'utf8',  # 编码类型
    'cursorclass': pymysql.cursors.DictCursor  # 按字典输出
}

db = pymysql.connect(**config)
cursor = db.cursor()

wb = load_workbook('F:\\comic.xlsx')
sheet = wb.worksheets[0]

# 遍历第一行，确定各属性位置
series_i = 'A'  # 同人作品原作名
title_i = 'B'  # 标题
author_i = 'C'  # 作者
publisher_i = 'D'  # 出版地
is_tankoubon_i = 'E'  # 是否是单行本
size_i = 'F'  # 大小
cover_i = 'G'  # 封面
desc_i = 'H'  # 描述
is_mushuusei_i = 'I'  # 是否无修
add_date_i = 'J'  # 添加日期
tag_i = 'K'  # 标签
rating_i = 'M'  # 评分（1~5）
lan_i = 'N'  # 语言
is_read_i = 'O'  # 是否已读
tran_title_i = 'P'  # 汉化名

nrow = sheet.max_row
for i in range(2, nrow + 1):
    comic = {
        'title': '',
        'authors': '',
        'tran_title': '',
        'cover_path': '',
        'tags': '',
        'description': '',
        'size': 0,
        'is_readed': False,
        'rating': 0,
        'is_mushuusei': False,
        'is_tankoubon': False,
        'is_color': False,
        'original_work': '',
        'create_time': '0000-00-00',
        'laguage': '中文',
        'publisher': '',
    }
    if i == 965:
        df = 34
    comic['title'] = sheet[title_i + str(i)].value
    if "'" in comic['title']:
        comic['title'] = comic['title'].replace("'", "\\'")
    comic['authors'] = str(sheet[author_i + str(i)].value)
    if "'" in comic['authors']:
        comic['authors'] = comic['authors'].replace("'", "\\'")
    comic['tran_title'] = sheet[tran_title_i + str(i)].value
    comic['cover_path'] = sheet[cover_i + str(i)].value
    if "'" in comic['cover_path']:
        comic['cover_path'] = comic['cover_path'].replace("'", "\\'")
    comic['tags'] = sheet[tag_i + str(i)].value
    comic['description'] = sheet[desc_i + str(i)].value
    if comic['description'] and "'" in comic['description']:
        comic['description'] = comic['description'].replace("'", "\\'")
    size = float(sheet[size_i + str(i)].value)
    size = size / 1024 / 1024
    size = round(size, 2)
    comic['size'] = size
    comic['is_readed'] = get_bool(sheet[is_read_i + str(i)].value)
    if sheet[rating_i + str(i)].value is None:
        comic['rating'] = 0
    else:
        comic['rating'] = int(sheet[rating_i + str(i)].value)
    comic['is_mushuusei'] = get_bool(sheet[is_mushuusei_i + str(i)].value)
    comic['is_tankoubon'] = get_bool(sheet[is_tankoubon_i + str(i)].value)
    if comic['tags'] and '全彩' in comic['tags']:
        comic['is_color'] = True
    comic['original_work'] = sheet[series_i + str(i)].value
    if sheet[lan_i + str(i)].value == 'jpn':
        comic['laguage'] = '日语'
    comic['publisher'] = sheet[publisher_i + str(i)].value
    comic['create_time'] = sheet[add_date_i + str(i)].value.replace('T', ' ').replace('+08:00', '')

    info = '[{}/{}] {}'.format(i, nrow, comic['title'])
    if comic['tran_title']:
        info += '/{}'.format(comic['tran_title'])
    print(info)

    sql = "select title from comic where title like '{}'".format(comic['title'])
    cursor.execute(sql)
    query = cursor.fetchone()
    if i == 1830:
        df = 34
    if not query:
        sql = "insert into comic(title,authors,tran_title,cover_path,tags,description,size,is_readed,rating," \
              "is_mushuusei,is_tankoubon,is_color,original_work,laguage,publisher,create_time) values(" \
              "'{}','{}','{}','{}','{}','{}',{},{},{},{},{},{},'{}','{}','{}','{}');".format(
            comic['title'], comic['authors'], comic['tran_title'], comic['cover_path'], comic['tags']
            , comic['description'], comic['size'], comic['is_readed'], comic['rating'], comic['is_mushuusei']
            , comic['is_tankoubon'], comic['is_color'], comic['original_work'], comic['laguage'], comic['publisher']
            , comic['create_time'])
        sql = sql.replace('None', '')
        cursor.execute(sql)
        db.commit()
cursor.close()
db.close()