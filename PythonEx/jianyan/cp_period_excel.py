#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
计算实际还款中每笔还款记录所对应的计划表中的期数（excel数据源）
"""

from openpyxl import load_workbook
import time

time_start = time.time()
time_last = time_start

wb_plan = load_workbook('F:\\还款计划表-0427.xlsx')
sheet_plan = wb_plan.worksheets[0]

time_now = time.time()
print('读取计划表耗时：{}s/{}s'.format(round(time_now - time_last, 2), round(time_now - time_start, 2)))
time_last = time_now

plans = {}
nrow = sheet_plan.max_row
for i in range(2, nrow + 1):
    tp_no = sheet_plan['I' + str(i)].value
    if not tp_no:
        continue
    # 判断本编号是否已建立词典
    if tp_no in plans:
        # 是则直接添加
        tp_plan = plans[tp_no]
        tp_plan.append(sheet_plan['H' + str(i)].value)
    else:
        # 否则新建编号记录列表再添加
        tp_plan = [sheet_plan['H' + str(i)].value]
        plans[tp_no] = tp_plan
    i += 1

time_now = time.time()
print('分类计划表耗时：{}s/{}s'.format(round(time_now - time_last, 4), round(time_now - time_start, 2)))
wb_plan.close()
time_last = time_now

# 获取所有实际还款，并建立词典

wb_real = load_workbook('F:\\实际还款plan.xlsx')
sheet_real = wb_real.worksheets[0]

time_now = time.time()
print('读取实际表耗时：{}s/{}s'.format(round(time_now - time_last, 2), round(time_now - time_start, 2)))
time_last = time_now

nos = []
reals = {}
nrow = sheet_real.max_row
for i in range(2, nrow + 1):
    tp_no = sheet_real['H' + str(i)].value
    if not tp_no in nos:
        nos.append(tp_no)
    tp_data = {
        'return_date': sheet_real['D' + str(i)].value,
        'money': sheet_real['E' + str(i)].value,
        'period_pos': 'F' + str(i),
        'name': sheet_real['G' + str(i)].value,
        'no': tp_no
    }
    # 判断本编号是否已建立词典
    if tp_no in reals:
        # 是则直接添加
        tp_real = reals[tp_no]
        tp_real.append(tp_data)
    else:
        # 否则新建编号记录列表再添加
        tp_real = [tp_data]
        reals[tp_no] = tp_real
    i += 1

time_now = time.time()
print('分类实际表耗时：{}s/{}s'.format(round(time_now - time_last, 4), round(time_now - time_start, 2)))
time_last = time_now

no_len = len(nos)
i = 1
for no in nos:
    # 获取该编号还款计划
    if not no in plans: # 跳过计划中无编号数据
        i += 1
        continue
    now_plans = plans[no]

    # 获取实际还款期数
    now_reals = reals[no]
    # 对其按时间升序（快速排序）
    for x in range(1, len(now_reals)):
        for j in range(x - 1, -1, -1):
            a = now_reals[j]['return_date']
            b = now_reals[j + 1]['return_date']
            if a > b:
                temp = now_reals[j]
                now_reals[j] = now_reals[j + 1]
                now_reals[j + 1] = temp
            else:
                break

    # 将计划与实际还款进行对比，计算其实际还款数量
    now_period = 0  # 当前还款期数，最后写入时加1
    now_money = 0  # 当前还款金额
    plan_money = now_plans[0]  # 当前应还款金额累计，与期数对应，初始为第1期
    for info in now_reals:
        # 在值不为空时叠加（反例：编号：HXYLcredit201706283794467）
        if info['money']:
            now_money += info['money']  # 叠加本次还款金额
        # 判断当前还款是否大于当前应还款
        if now_money > plan_money:
            # 大于时计算还至第几期
            while now_money > plan_money:
                # 判断是否已还至最后一期，至最后一期时不再往后计算
                if now_period == len(now_plans) - 1:
                    break
                now_period += 1
                plan_money += now_plans[now_period]
        time_now = time.time()
        print('[{}/{}]\t编号：{}\t姓名：{}\t还款时间：{}\t金额：{}/{}\t期数：{}\t位置：{}'.format(i, no_len, no,
                                                                              info['name'], info['return_date'],
                                                                              info['money'], now_plans[now_period],
                                                                              now_period + 1, info['period_pos']))
        # 写入
        pos = info['period_pos']
        sheet_real[pos].number_format = 'General'   # 设置单元格类型为常规，避免写入后出现小数点的情况
        sheet_real[pos] = now_period + 1
    time_now = time.time()
    print('本编号计算耗时：{}s/{}s'.format(round(time_now - time_last, 4), round(time_now - time_start, 2)))
    time_last = time_now
    i += 1

print('保存数据中……')
wb_real.save('F:\\test.xlsx')
wb_real.close()
print('计算完毕！，累计耗时{}s'.format(round(time.time() - time_start, 2)))

