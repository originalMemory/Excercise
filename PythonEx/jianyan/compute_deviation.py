#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
本文件用于对excel实验数据处理
作者：吴厚波
创建时间：2018-6-2
更新时间：2018-6-12
"""

from openpyxl import load_workbook
from openpyxl import Workbook

wb_source = load_workbook('F:\\M1-三天数据处理过程.xlsx')  # 总表
sheet_source = wb_source.worksheets[3]

wb_target = Workbook()
sheet_target = wb_target.get_active_sheet()

letter = ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V',
          'W', 'X', 'Y', 'Z')

def get_col_letter(col_index):
    """
    获取列数字序号对应的字母序号（仅至双字母编号，无法计算三字母编号）
    :param col_index: 列数字序号（从0开始）
    :return: 列字母序号
    """
    # 判断序号是否已超过AA
    if col_index <= 25:
        return letter[col_index]
    else:
        # 判断首字母序号
        initial = letter[int(col_index / 26) - 1]
        # 获取后续字母序号
        second = letter[col_index % 26]
        return initial + second


def get_col_index(col_letter):
    """
    获取列字母序号对应的数字序号（仅至双字母编号，无法计算三字母编号）
    :param col_letter: 列字母序号
    :return: 列数字序号（从0开始）
    """
    # 判断序号是否已超过AA
    if len(col_letter) == 1:
        return letter.index(col_letter)
    else:
        # 判断首字母序号
        index = (letter.index(col_letter[0]) + 1) * 26
        # 获取后续字母序号
        index += letter.index(col_letter[1])
        return index


def generate_new_dict(source_dict):
    """
    根据原key生成key误差范围在0.005以的新字典
    :param source_dict:原字典
    :return:新字典
    """
    tp_list = list(source_dict.keys())
    source_keys = list(map(lambda x: float(x), tp_list))
    used_index = []  # 已经计算过的key值，用以避免反复计算
    new_dict = {}
    for i in range(len(source_keys)):
        if i in used_index:
            continue
        new_key = source_keys[i]  # 新的key值
        num = 1  # 累加了几行
        new_value = {}  # 新key对应的value,也是dict
        new_value = source_dict[str(source_keys[i])]
        used_index.append(i)
        # 遍历余下行，寻找在误差范围内行
        for j in range(i + 1, len(source_keys)):
            if abs(source_keys[i] - source_keys[j]) <= deviation and not j in used_index:
                new_key += source_keys[j]
                num += 1
                used_index.append(j)
                # 累加该行的值到新行中
                for tp_key, tp_value in source_dict[str(source_keys[j])].items():
                    if tp_key in new_value:
                        new_value[tp_key] += tp_value
                    else:
                        new_value[tp_key] = tp_value
        new_key /= num  # 计算均值
        new_dict[str(new_key)] = new_value
    return new_dict


# 遍历表格，获取exp,blank和after的数据列表
col_exp = 'A'
col_blank = 'M'
col_after = 'X'
deviation = 0.005
nrows = sheet_source.max_row
ncol = sheet_source.max_column
# 原始数据字典
exps_source = {}
afters_source = {}
blanks_source = {}
# 排序后的字典
exps_target = {}
blanks_target = {}
afters_target = {}

# 遍历生成原始数据字典
col_exp_index = get_col_index(col_exp)
col_blank_index = get_col_index(col_blank)
col_after_index = get_col_index(col_after)
for i in range(2, nrows + 1):
    # 本行key值，如55.0551
    exp_key = str(sheet_source[col_exp + str(i)].value)
    if exp_key and not exp_key in exps_source and isinstance(sheet_source[col_exp + str(i)].value, float):
        exps_source[exp_key] = {}
    blank_key = str(sheet_source[col_blank + str(i)].value)
    if blank_key and not blank_key in blanks_source and isinstance(sheet_source[col_blank + str(i)].value, float):
        blanks_source[blank_key] = {}
    after_key = str(sheet_source[col_after + str(i)].value)
    if after_key and not after_key in afters_source and isinstance(sheet_source[col_after + str(i)].value, float):
        afters_source[after_key] = {}
    # 遍历该行所有列
    for j in range(ncol):
        # 跳过所有key值列
        if j == col_exp_index or j == col_blank_index or j == col_after_index:
            continue
        cell_pos = get_col_letter(j) + str(i)  # 本单元格位置
        if not sheet_source[cell_pos].value or not isinstance(sheet_source[cell_pos].value, float):
            continue

        col_key = sheet_source[get_col_letter(j) + '1'].value  # 本列key值，如20170621-m52347-09.18.RAW
        # exp区域
        if j < col_blank_index:
            # 判断是否本行关键词已存在且同一列的值也存在，存在则累加
            if col_key in exps_source[exp_key]:
                exps_source[exp_key][col_key] += sheet_source[cell_pos].value
            else:
                exps_source[exp_key][col_key] = sheet_source[cell_pos].value
        # blank区域
        elif j < col_after_index:
            # 判断是否本行关键词已存在且同一列的值也存在，存在则累加
            if col_key in blanks_source[blank_key]:
                blanks_source[blank_key][col_key] += sheet_source[cell_pos].value
            else:
                blanks_source[blank_key][col_key] = sheet_source[cell_pos].value
        # after区域
        else:
            # 判断是否本行关键词已存在且同一列的值也存在，存在则累加
            if col_key in afters_source[after_key]:
                afters_source[after_key][col_key] += sheet_source[cell_pos].value
            else:
                afters_source[after_key][col_key] = sheet_source[cell_pos].value

print('数据加载完毕！')
# 分别遍历原始数据字典key值，生成新的字典
exps_target = generate_new_dict(exps_source)
blanks_target = generate_new_dict(blanks_source)
afters_target = generate_new_dict(afters_source)
print('数据合并完毕！')
# 建立新表表头，同时建立列关键字list
exp_col_keys = []
sheet_target[col_exp + '1'] = 'Sample-exp'
for i in range(col_exp_index + 1, col_blank_index):
    cell_pos = get_col_letter(i) + '1'  # 本单元格位置
    if sheet_source[cell_pos].value:
        sheet_target[cell_pos] = sheet_source[cell_pos].value
        exp_col_keys.append(sheet_source[cell_pos].value)
    else:
        break

# 遍历新字典，计算新的exps的数据
i = 2
for exp_key, exp_value in exps_target.items():
    exp_num = float(exp_key)
    sheet_target[col_exp + str(i)] = exp_num
    tp_dict = {}  # 临时计算用数据字典
    num = 0  # 符合条件的行数
    # 遍历blanks，计算累加值
    for blank_key, blank_value in blanks_target.items():
        blank_num = float(blank_key)
        if abs(exp_num - blank_num) <= deviation:
            # 遍历累加本行数据
            for tp_key, tp_value in blank_value.items():
                if not tp_key in tp_dict:
                    tp_dict[tp_key] = tp_value
                else:
                    tp_dict[tp_key] += tp_value
            num += 1
    # 遍历after，计算累加值
    for after_key, after_value in afters_target.items():
        after_num = float(after_key)
        if abs(exp_num - after_num) <= deviation:
            # 遍历累加本行数据
            for tp_key, tp_value in after_value.items():
                if not tp_key in tp_dict:
                    tp_dict[tp_key] = tp_value
                else:
                    tp_dict[tp_key] += tp_value
            num += 1

    # 计算数据均值
    for tp_key, tp_value in tp_dict.items():
        tp_dict[tp_key] = tp_value / num

    # 计算最后差值结果
    for j in range(len(exp_col_keys)):
        col_letter = get_col_letter(col_exp_index + 1 + j)
        col_key=exp_col_keys[j]
        result=0.0
        if col_key in exp_value:
            if col_key in tp_dict:
                result= exp_value[col_key] - tp_dict[col_key]
            else:
                result = exp_value[col_key]
        else:
            if col_key in tp_dict:
                result = 0 - tp_dict[col_key]
        sheet_target[col_letter + str(i)]=result
    i+=1


wb_target.save(r'F:\result.xlsx')
